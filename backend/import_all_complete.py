"""
Complete import of ALL 5 supplier catalogs with proper parsing
"""
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv
from pathlib import Path
import uuid
from datetime import datetime, timezone
import pandas as pd
import requests
from openpyxl import load_workbook
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

def download_file(url, filename):
    print(f"📥 Downloading {filename}...")
    r = requests.get(url)
    with open(filename, 'wb') as f:
        f.write(r.content)
    return filename

def parse_romax(filename):
    """Special parser for Ромакс with merged cells"""
    print(f"📊 Parsing Ромакс (special format)...")
    
    wb = load_workbook(filename)
    sheet = wb.active
    
    products = []
    
    for row in sheet.iter_rows(min_row=11, values_only=True):  # Data starts at row 11
        try:
            # Column 0: Product name (includes article)
            # Column 7: Unit
            # Column 9: Price
            product_name = row[0]
            unit = row[7] if row[7] else 'кг'
            price_str = row[9]
            
            if not product_name or not price_str:
                continue
            
            # Clean price: "1 793,40" -> 1793.40
            price_clean = str(price_str).replace(' ', '').replace(',', '.')
            try:
                price = float(price_clean)
                if price <= 0 or price > 100000:
                    continue
            except:
                continue
            
            # Extract article from product name if it has №
            article_match = re.search(r'№(\d+)', str(product_name))
            article = article_match.group(1) if article_match else str(len(products))
            
            products.append({
                'productName': str(product_name)[:200].strip(),
                'price': price,
                'article': article,
                'unit': str(unit).strip() if unit else 'кг'
            })
            
        except Exception as e:
            continue
    
    print(f"✅ Parsed {len(products)} products from Ромакс")
    return products

def parse_standard(filename, supplier_name):
    """Standard parser for most files"""
    print(f"📊 Parsing {supplier_name}...")
    
    try:
        xl = pd.ExcelFile(filename)
        all_products = []
        
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
            
            # Find header
            header_row = -1
            name_col = None
            price_col = None
            unit_col = None
            article_col = None
            
            for idx, row in df.iterrows():
                if idx > 20:
                    break
                
                row_str = ' '.join([str(x).lower() for x in row if pd.notna(x)])
                
                if any(k in row_str for k in ['наименование', 'название', 'товар', 'name']):
                    header_row = idx
                    
                    for col_idx, cell in enumerate(row):
                        if pd.isna(cell):
                            continue
                        cell_str = str(cell).lower()
                        
                        if 'наименование' in cell_str or 'название' in cell_str:
                            name_col = col_idx
                        elif 'цена' in cell_str or 'price' in cell_str:
                            price_col = col_idx
                        elif 'ед.' in cell_str or 'unit' in cell_str:
                            unit_col = col_idx
                        elif 'артикул' in cell_str or 'код' in cell_str:
                            article_col = col_idx
                    
                    if name_col is not None and price_col is not None:
                        break
            
            # Parse data
            start_row = header_row + 1 if header_row >= 0 else 0
            
            for idx, row in df.iterrows():
                if idx < start_row:
                    continue
                
                name = row[name_col] if name_col is not None and name_col < len(row) else row[0] if len(row) > 0 else None
                if pd.isna(name) or str(name).strip() == '':
                    continue
                
                price = row[price_col] if price_col is not None and price_col < len(row) else row[1] if len(row) > 1 else None
                if pd.isna(price):
                    continue
                
                try:
                    price_val = float(price)
                    if price_val <= 0:
                        continue
                except:
                    continue
                
                unit = 'шт'
                if unit_col is not None and unit_col < len(row) and not pd.isna(row[unit_col]):
                    unit = str(row[unit_col]).strip()
                
                article = str(idx)
                if article_col is not None and article_col < len(row) and not pd.isna(row[article_col]):
                    article = str(row[article_col]).strip()
                
                all_products.append({
                    'productName': str(name).strip()[:200],
                    'price': price_val,
                    'article': article,
                    'unit': unit
                })
        
        print(f"✅ Parsed {len(all_products)} products from {supplier_name}")
        return all_products
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return []

async def import_all_complete():
    print("🔄 COMPLETE IMPORT OF ALL 5 CATALOGS")
    print("=" * 70)
    
    suppliers_data = [
        {
            'name': 'Алиди',
            'url': 'https://customer-assets.emergentagent.com/job_resto-supplier/artifacts/m287cc6h_%D0%90%D0%BB%D0%B8%D0%B4%D0%B8%20%281%29.xlsx',
            'parser': 'standard'
        },
        {
            'name': 'ВЗ',
            'url': 'https://customer-assets.emergentagent.com/job_resto-supplier/artifacts/49b50io7_%D0%92%D0%97.xlsx',
            'parser': 'standard'
        },
        {
            'name': 'ПраймФудс',
            'url': 'https://customer-assets.emergentagent.com/job_resto-supplier/artifacts/0v37whfj_%D0%9F%D1%80%D0%B0%D0%B8%CC%86%D0%BC%D0%A4%D1%83%D0%B4%D1%81.xlsx',
            'parser': 'standard'
        },
        {
            'name': 'Ромакс',
            'url': 'https://customer-assets.emergentagent.com/job_resto-supplier/artifacts/dg7tz9ss_%D0%A0%D0%BE%D0%BC%D0%B0%D0%BA%D1%81%20%281%29.xlsx',
            'parser': 'romax'
        },
        {
            'name': 'VICI',
            'url': 'https://customer-assets.emergentagent.com/job_resto-supplier/artifacts/fsmj7tfk_VICI%20.xlsx',
            'parser': 'standard'
        }
    ]
    
    # Clear all price lists
    print("\n🗑️  Clearing all existing products...")
    result = await db.price_lists.delete_many({})
    print(f"   Deleted {result.deleted_count} products")
    
    total_imported = 0
    
    for supplier_data in suppliers_data:
        print(f"\n{'='*70}")
        print(f"📦 {supplier_data['name']}")
        print(f"{'='*70}")
        
        # Get company
        company = await db.companies.find_one({"companyName": supplier_data['name']}, {"_id": 0})
        if not company:
            print(f"❌ Company not found, skipping...")
            continue
        
        # Download
        filename = f"/tmp/{supplier_data['name']}.xlsx"
        download_file(supplier_data['url'], filename)
        
        # Parse
        if supplier_data['parser'] == 'romax':
            products = parse_romax(filename)
        else:
            products = parse_standard(filename, supplier_data['name'])
        
        if not products:
            print(f"⚠️  No products parsed, skipping...")
            continue
        
        # Import
        print(f"💾 Importing {len(products)} products...")
        for i, product in enumerate(products):
            await db.price_lists.insert_one({
                'id': str(uuid.uuid4()),
                'supplierCompanyId': company['id'],
                'productName': product['productName'],
                'article': product['article'],
                'price': product['price'],
                'unit': product['unit'],
                'availability': True,
                'active': True,
                'createdAt': datetime.now(timezone.utc).isoformat(),
                'updatedAt': datetime.now(timezone.utc).isoformat()
            })
            
            if (i + 1) % 200 == 0:
                print(f"  Imported {i + 1}/{len(products)}...")
        
        print(f"✅ Completed: {len(products)} products")
        total_imported += len(products)
    
    print(f"\n{'='*70}")
    print(f"🎉 IMPORT COMPLETE")
    print(f"{'='*70}")
    
    # Verify
    print(f"\n📋 Final verification:")
    suppliers = await db.companies.find({"type": "supplier"}, {"_id": 0, "companyName": 1, "id": 1}).to_list(20)
    grand_total = 0
    
    for s in suppliers:
        count = await db.price_lists.count_documents({"supplierCompanyId": s['id']})
        status = "✅" if count > 50 else ("⚠️" if count > 0 else "❌")
        print(f"  {status} {s['companyName']}: {count} products")
        grand_total += count
    
    print(f"\n📊 GRAND TOTAL: {grand_total} products across {len(suppliers)} suppliers")
    
    client.close()

if __name__ == "__main__":
    asyncio.run(import_all_complete())
