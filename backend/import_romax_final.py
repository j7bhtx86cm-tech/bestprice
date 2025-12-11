import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv
from pathlib import Path
import uuid
from datetime import datetime, timezone
from openpyxl import load_workbook
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

async def import_romax():
    client = AsyncIOMotorClient(os.environ['MONGO_URL'])
    db = client[os.environ['DB_NAME']]
    
    romax = await db.companies.find_one({"companyName": "Ромакс"}, {"_id": 0})
    if not romax:
        print("❌ Ромакс not found")
        return
    
    print("📦 Importing Ромакс...")
    
    wb = load_workbook('/tmp/new_catalog.xlsx')
    ws = wb['Ромакс']
    
    products = []
    row_num = 0
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        row_num += 1
        try:
            if not row[2]:
                continue
            
            name = str(row[2]).strip()
            unit = str(row[3]).strip() if row[3] else 'КОР'
            min_qty = int(row[4]) if row[4] else 1
            
            price_val = row[5]
            if isinstance(price_val, (int, float)):
                price = float(price_val)
            else:
                price = float(str(price_val).replace(' ', '').replace(',', '.'))
            
            if price <= 0 or price > 1000000:
                continue
            
            article = str(row_num)
            match = re.search(r'№(\d+)', name)
            if match:
                article = match.group(1)
            
            products.append({
                'name': name[:200],
                'article': article,
                'price': price,
                'unit': unit,
                'minQty': min_qty
            })
            
        except:
            continue
    
    print(f"✅ Parsed {len(products)} products")
    
    if products:
        print(f"💾 Importing...")
        for i, p in enumerate(products):
            await db.price_lists.insert_one({
                'id': str(uuid.uuid4()),
                'supplierCompanyId': romax['id'],
                'productName': p['name'],
                'article': p['article'],
                'price': p['price'],
                'unit': p['unit'],
                'minQuantity': p['minQty'],
                'availability': True,
                'active': True,
                'createdAt': datetime.now(timezone.utc).isoformat(),
                'updatedAt': datetime.now(timezone.utc).isoformat()
            })
            
            if (i + 1) % 200 == 0:
                print(f"  {i + 1}/{len(products)}...")
    
    count = await db.price_lists.count_documents({"supplierCompanyId": romax['id']})
    print(f"\n✅ Ромакс: {count} products")
    
    total = await db.price_lists.count_documents({})
    print(f"📊 GRAND TOTAL: {total} products")
    
    client.close()

asyncio.run(import_romax())
