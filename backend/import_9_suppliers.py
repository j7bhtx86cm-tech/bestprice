"""
Import new catalog with 9 suppliers
"""
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv
from pathlib import Path
import uuid
from datetime import datetime, timezone
import pandas as pd
import bcrypt
from openpyxl import load_workbook
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

async def create_supplier(name, email):
    company = await db.companies.find_one({"companyName": name}, {"_id": 0})
    if company:
        return company['id']
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        user_id = str(uuid.uuid4())
        await db.users.insert_one({
            'id': user_id,
            'email': email,
            'passwordHash': hash_password('password123'),
            'role': 'supplier',
            'createdAt': datetime.now(timezone.utc).isoformat(),
            'updatedAt': datetime.now(timezone.utc).isoformat()
        })
    else:
        user_id = user['id']
    
    company_id = str(uuid.uuid4())
    await db.companies.insert_one({
        'id': company_id,
        'type': 'supplier',
        'userId': user_id,
        'inn': f"{7700000000 + abs(hash(name)) % 1000000}",
        'ogrn': f"{1027700000000 + abs(hash(name)) % 1000000}",
        'companyName': name,
        'legalAddress': f"г. Москва, {name}",
        'actualAddress': f"г. Москва, {name}",
        'phone': '+7 (495) 000-00-00',
        'email': email,
        'contractAccepted': True,
        'createdAt': datetime.now(timezone.utc).isoformat(),
        'updatedAt': datetime.now(timezone.utc).isoformat()
    })
    
    return company_id

async def import_all():
    print("🔄 IMPORTING 9 SUPPLIERS")
    print("=" * 70)
    
    # Clear old data
    print("\n🗑️  Clearing old data...")
    await db.price_lists.delete_many({})
    old_suppliers = await db.companies.find({"type": "supplier"}, {"_id": 0}).to_list(100)
    for s in old_suppliers:
        await db.users.delete_many({"id": s['userId']})
        await db.companies.delete_one({"id": s['id']})
    
    # Read file
    filename = '/tmp/new_catalog.xlsx'
    xl = pd.ExcelFile(filename)
    
    supplier_emails = {
        'Алиди': 'alidi@bestprice.ru',
        'Айфрут': 'ayfruit@bestprice.ru',
        'Фаворит': 'favorit@bestprice.ru',
        'ТД ДУНАЙ': 'dunay@bestprice.ru',
        'Интегрита': 'integrita@bestprice.ru',
        'Ромакс': 'romax@bestprice.ru',
        'Прайфуд': 'primefood@bestprice.ru',
        'Vici': 'vici@bestprice.ru',
        'В-З': 'vz@bestprice.ru'
    }
    
    total = 0
    
    for sheet_name in xl.sheet_names:
        print(f"\n{'='*70}")
        print(f"📦 {sheet_name}")
        print(f"{'='*70}")
        
        company_id = await create_supplier(sheet_name, supplier_emails.get(sheet_name, f"{sheet_name.lower()}@bestprice.ru"))
        
        # Handle Ромакс separately (different format)
        if sheet_name == 'Ромакс':
            wb = load_workbook(filename)
            ws = wb[sheet_name]
            
            products = []
            for row in ws.iter_rows(min_row=4, values_only=True):
                try:
                    if not row[2]: continue
                    name = str(row[2]).strip()
                    unit = str(row[3]).strip() if row[3] else 'КОР'
                    min_qty = int(row[4]) if row[4] else 1
                    
                    price_val = row[5]
                    if isinstance(price_val, (int, float)):
                        price = float(price_val)
                    else:
                        price = float(str(price_val).replace(' ', '').replace(',', '.'))
                    
                    if price <= 0: continue
                    
                    article = str(row_num) if not row[1] else str(row[1])
                    match = re.search(r'№(\d+)', name)
                    if match: article = match.group(1)
                    
                    products.append({
                        'name': name[:200],
                        'article': article,
                        'price': price,
                        'unit': unit,
                        'minQty': min_qty
                    })
                except:
                    continue
        else:
            # Standard format
            df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
            products = []
            
            for idx in range(2, len(df)):
                row = df.iloc[idx]
                try:
                    if pd.isna(row[2]): continue
                    
                    code = str(row[1]).strip() if pd.notna(row[1]) else str(idx)
                    name = str(row[2]).strip()
                    unit = str(row[3]).strip() if pd.notna(row[3]) else 'шт'
                    min_qty = int(row[4]) if pd.notna(row[4]) else 1
                    price = float(row[5]) if pd.notna(row[5]) else 0
                    
                    if price <= 0: continue
                    
                    products.append({
                        'name': name[:200],
                        'article': code,
                        'price': price,
                        'unit': unit,
                        'minQty': min_qty
                    })
                except:
                    continue
        
        if products:
            print(f"💾 Importing {len(products)} products...")
            for i, p in enumerate(products):
                await db.price_lists.insert_one({
                    'id': str(uuid.uuid4()),
                    'supplierCompanyId': company_id,
                    'productName': p['name'],
                    'article': p['article'],
                    'price': p['price'],
                    'unit': p['unit'],
                    'minQuantity': p['minQty'],
                    'availability': True,
                    'active': True,
                    'createdAt': datetime.now(timezone.utc).isoformat(),
                    'updatedAt': datetime.now(timezone.utc).isoformat()
                })
                
                if (i + 1) % 300 == 0:
                    print(f"  {i + 1}/{len(products)}...")
            
            print(f"✅ {len(products)} products")
            total += len(products)
        else:
            print(f"⚠️  No products")
    
    print(f"\n{'='*70}")
    print(f"🎉 Total: {total} products")
    print(f"{'='*70}")
    
    # Verify
    suppliers = await db.companies.find({"type": "supplier"}, {"_id": 0, "companyName": 1, "id": 1}).to_list(20)
    for s in suppliers:
        cnt = await db.price_lists.count_documents({"supplierCompanyId": s['id']})
        print(f"  {s['companyName']}: {cnt}")
    
    client.close()

if __name__ == "__main__":
    asyncio.run(import_all())
