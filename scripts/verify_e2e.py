#!/usr/bin/env python3
"""
E2E verification: supplier login, settings, pricelist upload/import, catalog visibility.
Requires: backend running, init_suppliers.py already run (Excel with valid passwords).
Exits 0 with "✅ ALL CHECKS PASSED" or 1 with error message.
Do not use ripgrep (rg) — run on Mac without it.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "_exports" / "supplier_accesses.xlsx"
FIXTURE_CSV = ROOT / "tests" / "fixtures" / "pricelist_sample.csv"


def main() -> int:
    base_url = os.environ.get("VERIFY_BASE_URL", "http://127.0.0.1:8000").rstrip("/")
    api = f"{base_url}/api"

    # 1) Excel exists
    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found. Run: python scripts/init_suppliers.py")
        return 1

    # 2) Read first rows from Excel (need openpyxl)
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl required. Install: pip install openpyxl")
        return 1
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_row = tuple(h for h in header_row if h is not None)
    idx = {str(h): i for i, h in enumerate(header_row)}
    data_rows = list(ws.iter_rows(min_row=2, max_row=4, values_only=True))
    wb.close()

    def get(row, key):
        i = idx.get(key)
        if i is None:
            return None
        row = tuple(row)
        return row[i] if i < len(row) else None

    # Pick first row with a usable password (not "unchanged")
    email, password_plain, company_id = None, None, None
    for row in data_rows:
        if not row:
            continue
        em = get(row, "supplier_email")
        pw = get(row, "password_plain")
        cid = get(row, "company_id")
        if em and cid and pw and "unchanged" not in (pw or ""):
            email, password_plain, company_id = str(em), str(pw), str(cid)
            break
    if not email or not password_plain or not company_id:
        print("ERROR: No row in Excel with valid password (run init_suppliers on clean DB to generate passwords)")
        return 1

    try:
        import requests
    except ImportError:
        print("ERROR: requests required. Install: pip install requests")
        return 1

    def req(method, path, **kwargs):
        r = requests.request(method, f"{api}{path}", timeout=30, **kwargs)
        return r

    # 3) Login
    r = req("POST", "/auth/login", json={"email": email, "password": password_plain})
    if r.status_code != 200:
        print(f"ERROR: Login failed: {r.status_code} {r.text}")
        return 1
    token = r.json().get("access_token")
    if not token:
        print("ERROR: No access_token in login response")
        return 1
    headers_auth = {"Authorization": f"Bearer {token}"}

    # 4) GET supplier-settings/my
    r = req("GET", "/supplier-settings/my", headers=headers_auth)
    if r.status_code != 200:
        print(f"ERROR: GET supplier-settings/my: {r.status_code} {r.text}")
        return 1

    # 5) PUT supplier-settings/my (change one field), then GET and check
    r = req("PUT", "/supplier-settings/my", headers=headers_auth, json={"minOrderAmount": 9999})
    if r.status_code != 200:
        print(f"ERROR: PUT supplier-settings/my: {r.status_code} {r.text}")
        return 1
    r = req("GET", "/supplier-settings/my", headers=headers_auth)
    if r.status_code != 200:
        print(f"ERROR: GET supplier-settings/my after PUT: {r.status_code}")
        return 1
    if r.json().get("minOrderAmount") != 9999:
        print(f"ERROR: Settings not updated: minOrderAmount = {r.json().get('minOrderAmount')}")
        return 1
    # Restore so we don't leave test data
    req("PUT", "/supplier-settings/my", headers=headers_auth, json={"minOrderAmount": 0})

    # 6) Upload fixture
    if not FIXTURE_CSV.exists():
        print(f"ERROR: Fixture not found: {FIXTURE_CSV}")
        return 1
    with open(FIXTURE_CSV, "rb") as f:
        files = {"file": ("pricelist_sample.csv", f, "text/csv")}
        r = req("POST", "/price-lists/upload", headers=headers_auth, files=files)
    if r.status_code not in (200, 201):
        print(f"ERROR: Upload failed: {r.status_code} {r.text}")
        return 1
    # Fixture has columns: productName, article, price, unit
    mapping = {"productName": "productName", "article": "article", "price": "price", "unit": "unit"}

    # 7) Import
    with open(FIXTURE_CSV, "rb") as f:
        files = {"file": ("pricelist_sample.csv", f, "text/csv")}
        data = {"column_mapping": __import__("json").dumps(mapping)}
        r = req("POST", "/price-lists/import", headers=headers_auth, files=files, data=data)
    if r.status_code not in (200, 201):
        print(f"ERROR: Import failed: {r.status_code} {r.text}")
        return 1

    # 8) GET price-lists/my >= 1
    r = req("GET", "/price-lists/my", headers=headers_auth)
    if r.status_code != 200:
        print(f"ERROR: GET price-lists/my: {r.status_code} {r.text}")
        return 1
    items = r.json() if isinstance(r.json(), list) else []
    if len(items) < 1:
        print(f"ERROR: GET price-lists/my returned {len(items)} items, expected >= 1")
        return 1

    # 9) GET suppliers/{company_id}/price-lists >= 1
    r = req("GET", f"/suppliers/{company_id}/price-lists")
    if r.status_code != 200:
        print(f"ERROR: GET suppliers/{{id}}/price-lists: {r.status_code} {r.text}")
        return 1
    catalog = r.json() if isinstance(r.json(), list) else []
    if len(catalog) < 1:
        print(f"ERROR: GET suppliers/{{id}}/price-lists returned {len(catalog)} items, expected >= 1")
        return 1

    print("✅ ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
